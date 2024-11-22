import sqlite3 as sq
import os
import pandas as pd
import openpyxl as xl
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

"""
Converts Excel File into database for easy SQL queries.
Used for v1-v6 later Enterprise Data Warehouse queried instead of Excel File
"""
def generate_ADE_DB(path_xl, path_db):
    path_db = path_db + r'\database' + r'\ADE WD.db'
    try:
        # Read the Excel file, skipping the first row and setting headers to the second row
        # NOTE: Important to skip first row as Workday exports it with title there.
        df = pd.read_excel(path_xl, header=1)


        # Create a connection to the SQLite database
        conn = sq.connect(path_db)
        cursor = conn.cursor()

        # Keep the table name as 'ADE WD'
        table_name = 'ADE WD'
        
        # Convert the DataFrame to SQL, using the fixed table name
        df.to_sql(table_name, conn, if_exists='replace', index=False)

        # Commit changes and close the connection
        conn.commit()
        print(f"Data from {path_xl} has been successfully imported into {path_db} as table '{table_name}'.")

    except Exception as e:
        print(f"An error occurred: {e}")
    
    finally:
        if conn:
            conn.close()

def export_xl(path, path_export):
    import os

    file_name = 'FY 2025 PEOPLE.xlsx'
    output_path = os.path.join(path_export, file_name)

    try:
        # Establish database connection
        conn_ADE = sq.connect(path)

        # Query the database
        query = """
        SELECT *, [Current FTE]*[TOTAL FULL TIME BASE SALARY] AS [X-Comp]
        FROM
                (SELECT
                    Worker,
                    [EMP ID],
                    [Job Profile],
                    [Pay Rate Type],
                    [Position Start Date],
                    [Position End Date],
                    SUM(CASE 
                            WHEN [Base Pay Projected Distribution Amount] > 0 
                            THEN [Distribution FTE] 
                            ELSE 0  
                        END) AS [Current FTE],
                    (CASE 
                            WHEN [Job Profile] LIKE '%Student Assistant%'
                            THEN [Base Pay Projected Distribution Amount] * 800
                            ELSE ([General Salary Plan Amount (Full Time Rate)] * 12)
                        END) AS [TOTAL FULL TIME BASE SALARY],
                    (CASE
                        WHEN [Allowance Plan] = 'UWP - Practice Plan'
                        THEN [Annualized Amount]
                        ELSE 0
                    END) AS [Y-Comp]
                FROM [ADE WD]
                GROUP BY Worker, [EMP ID], [Job Profile], [Pay Rate Type]
                HAVING [TOTAL FULL TIME BASE SALARY] > 0) AS JOE
        """
        df = pd.read_sql_query(query, conn_ADE)

        # Export to Excel
        df.to_excel(output_path, index=False)
        print(f"Excel file exported: {file_name}")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Ensure the connection is closed
        if 'conn_ADE' in locals():
            conn_ADE.close()



def add_buckets(path_xl, path_ADE):
    try:
        # Establish database connection
        conn_ADE = sq.connect(path_ADE)
        
        # Query for distinct BUCKETS and their corresponding Distribution FTE values
        query = """
        SELECT
            ade.Worker,
            wt.BUCKET AS BUCKET,
            ade.[Distribution Percent]
        FROM
            [ADE WD] AS ade
        LEFT JOIN
            [updated bucketsv2] AS wt
        ON
            (wt.Activity = ade.Activity OR (wt.Activity IS NULL AND ade.Activity IS NULL))
            AND (wt.AllocationCostCenterID = ade.[Allocation Cost Center ID] OR (wt.AllocationCostCenterID IS NULL AND ade.[Allocation Cost Center ID] IS NULL))
            AND (wt.Function = ade.Function OR (wt.Function IS NULL AND ade.Function IS NULL))
            AND (wt.Fund = ade.Fund OR (wt.Fund IS NULL AND ade.Fund IS NULL))
            AND (wt.Gift = ade.Gift OR (wt.Gift IS NULL AND ade.Gift IS NULL))
            AND (wt.Grant = ade.Grant OR (wt.Grant IS NULL AND ade.Grant IS NULL))
            AND (wt.Program = ade.Program OR (wt.Program IS NULL AND ade.Program IS NULL))
            AND (wt.Resource = ade.Resource OR (wt.Resource IS NULL AND ade.Resource IS NULL))
        WHERE
            ade.[Base Pay Projected Distribution Amount] > 0;
        """
        df_buckets = pd.read_sql_query(query, conn_ADE)
        
        query_general = """
        SELECT *, [Current FTE]*[TOTAL FULL TIME BASE SALARY] AS [X-Comp]
        FROM
                (SELECT
                    Worker,
                    [EMP ID],
                    [Job Profile],
                    [Pay Rate Type],
                    [Position Start Date],
                    [Position End Date],
                    SUM(CASE 
                            WHEN [Base Pay Projected Distribution Amount] > 0 
                            THEN [Distribution FTE] 
                            ELSE 0  
                        END) AS [Current FTE],
                    (CASE 
                            WHEN [Job Profile] LIKE '%Student Assistant%'
                            THEN [Base Pay Projected Distribution Amount] * 800
                            ELSE ([General Salary Plan Amount (Full Time Rate)] * 12)
                        END) AS [TOTAL FULL TIME BASE SALARY],
                    (CASE
                        WHEN [Allowance Plan] = 'UWP - Practice Plan'
                        THEN [Annualized Amount]
                        ELSE 0
                    END) AS [Y-Comp]
                FROM [ADE WD]
                GROUP BY Worker, [EMP ID], [Job Profile], [Pay Rate Type]
                HAVING [TOTAL FULL TIME BASE SALARY] > 0) AS JOE
        """
        df_existing = pd.read_sql_query(query_general, conn_ADE)

        
        # Get distinct BUCKETS and add columns for each BUCKET
        buckets = df_buckets['BUCKET'].unique()
        for bucket in buckets:
            # Initialize new columns as float to handle potential fractional values
            df_existing[bucket] = 0.0
        
        # Update DataFrame with Distribution FTE values
        for _, row in df_buckets.iterrows():
            worker = row['Worker']
            bucket = row['BUCKET']
            total_fte = float(row['Distribution Percent'])
            df_existing.loc[df_existing['Worker'] == worker, bucket] = total_fte

        # Duplicate the bucket columns
        for bucket in buckets:
            new_column_name = f"{bucket}_AS_PERCENTS"
            df_existing[new_column_name] = df_existing[bucket]
        
        for _, row in df_buckets.iterrows():
            worker = row['Worker']
            bucket = row['BUCKET']
            total_fte = float(row['Distribution Percent'])
            current_salary = df_existing.loc[df_existing['Worker'] == worker, 'X-Comp'].values
            if len(current_salary) > 0:
                current_salary = current_salary[0]
                df_existing.loc[df_existing['Worker'] == worker, bucket] = total_fte * current_salary

        # Write the updated DataFrame directly to the target Excel file
        df_existing.to_excel(path_xl, index=False)
        
        # Open the Excel file for formatting
        wb = xl.load_workbook(path_xl)
        ws = wb.active
        
        # Define styles
        aptos_narrow_font = Font(name='Aptos Narrow')
        bold_style = NamedStyle(name='bold_style', font=Font(name='Aptos Narrow', bold=True))
        percent_style = NamedStyle(name='percent_style', number_format='0%', font=aptos_narrow_font)
        blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # Blue, Accent 1, Lighter 60%
        orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Orange, Accent 6, Lighter 60%
        purple_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")  # Purple, Accent 4, Lighter 60%


        # Register styles
        if 'bold_style' not in wb.named_styles:
            wb.add_named_style(bold_style)
        if 'percent_style' not in wb.named_styles:
            wb.add_named_style(percent_style)

        # Apply bold style to headers
        for cell in ws[1]:
            cell.style = bold_style
        
        # Apply percentage formatting to the duplicated percent columns
        for bucket in buckets:
            col_idx = df_existing.columns.get_loc(bucket) + 1
            ws.cell(row=1, column=col_idx).fill = blue_fill
            col_name_percent = f"{bucket}_AS_PERCENTS"
            col_idx_percent = df_existing.columns.get_loc(col_name_percent) + 1
            ws.cell(row=1, column=col_idx_percent).fill = orange_fill
            column = f"{bucket}_AS_PERCENTS"
            if column in df_existing.columns:
                col_index = df_existing.columns.get_loc(column) + 1
                for row in ws.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
                    for cell in row:
                        cell.style = percent_style
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Define fills
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Blue, Accent 1, Lighter 60%
        orange_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Orange, Accent 6, Lighter 60%
        purple_fill = PatternFill(start_color="c7b1e0", end_color="c7b1e0", fill_type="solid")  # Purple, Accent 4, Lighter 60%

        # Apply fills to headers
        for col_idx, col_name in enumerate(df_existing.columns, start=1):
            if col_name in buckets:
                # Apply blue fill for BUCKET columns
                ws.cell(row=1, column=col_idx).fill = blue_fill
            elif col_name.endswith("_AS_PERCENTS"):
                # Apply orange fill for BUCKET_AS_PERCENTS columns
                ws.cell(row=1, column=col_idx).fill = orange_fill
            else:
                # Apply purple fill for other headers
                ws.cell(row=1, column=col_idx).fill = purple_fill


        # Save the changes back to the same Excel file
        wb.save(path_xl)
        
        print(f"Excel file updated with duplicated BUCKETS columns and saved to: {path_xl}")

    except Exception as e:
        print(f"An error occurred: {e}")
    
    finally:
        if 'conn_ADE' in locals():
            conn_ADE.close()
